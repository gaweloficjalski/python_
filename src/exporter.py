#Eksport do Excela
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def write_section (ws,title,data,title2):
    ws.append([title, title2])
    for city, number in data.items():
        ws.append([city,number])
    ws.append([])
    

def export_to_excel(df, results, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    # Podsumowanie
    ws.title = "Summary"
    ws['A1'] = "City"
    ws['B1'] = "Number of offers"
    for miasto,liczba in results['top_cities'].items():
        ws.append([miasto,liczba])
    ws.append([])
    
    # Zarobki
    ws2 = wb.create_sheet("Salaries")
    write_section(ws2, "Avg salary per city", results['avg_salary'],"Avg_salary")
    write_section(ws2, "Avg salary per experience", results['avg_salary_per_experience'],"Avg_salary")
    write_section(ws2, "Avg salary per category", results['avg_salary_per_categories'], "Avg_salary")

    # Skills
    ws3 = wb.create_sheet("Skills")
    write_section(ws3, "Top skills", results['top_skills'], "Number")
    
    # Raw Data
    columns = ['title', 'workplaceType', 'workingTime', 'experienceLevel','city','companyName', 'category_name','salary_from','salary_to', 'salary_avg', 'skills' ]
    df_clean = df[columns]
    ws4 = wb.create_sheet("Raw Data")
    for row in dataframe_to_rows(df_clean, index=False, header=True):
        ws4.append(row)
    
    for sheet in [ws, ws2, ws3, ws4]:
        _header_style(sheet, 1, sheet.max_column)
        _auto_col_width(sheet, 10, 30)
        _zebra_rows(sheet, 2, sheet.max_row, sheet.max_column)
        _thin_border(sheet, 1, sheet.max_row, sheet.max_column)
    wb.save(filepath)

NAVY       = "1B2A4A"
ACCENT     = "0EA5E9"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EFF6FF"
MID_GRAY   = "CBD5E1"

def _header_style(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill  = PatternFill("solid", fgColor=NAVY)
        cell.font  = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_col_width(ws, min_w: int = 10, max_w: int = 40) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)
        
def _zebra_rows(ws, data_start: int, data_end: int, col_count: int) -> None:
    for row in range(data_start, data_end + 1):
        fill = PatternFill("solid", fgColor=LIGHT_BLUE if row % 2 == 0 else WHITE)
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).fill = fill


def _thin_border(ws, row_start: int, row_end: int, col_count: int) -> None:
    thin = Side(style="thin", color=MID_GRAY)
    for row in range(row_start, row_end + 1):
        for col in range(1, col_count + 1):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                left=thin, right=thin,
                top=existing.top, bottom=existing.bottom,
            )
